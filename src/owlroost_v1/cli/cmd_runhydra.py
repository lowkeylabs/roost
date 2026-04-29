# src/owlroost/cli/cmd_run.py

from pathlib import Path

import click
from hydra import compose, initialize_config_dir
from loguru import logger
from openpyxl import load_workbook

from owlroost.core.override_parser import hydra_overrides_to_dict
from owlroost.core.owl_runner import run_single_case

# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------


def load_hydra_cfg(overrides=None):
    """
    Load Hydra configuration from ./conf relative to the current working directory.

    Supports SINGLE-RUN overrides only.
    """
    conf_dir = Path.cwd() / "conf"

    if not conf_dir.exists():
        raise RuntimeError(
            f"Hydra config directory not found: {conf_dir}\n"
            "Expected to run from a directory containing ./conf"
        )

    with initialize_config_dir(
        config_dir=str(conf_dir),
        version_base=None,
    ):
        return compose(
            config_name="config",
            overrides=overrides or [],
        )


def reject_sweeps(overrides):
    """
    Explicitly reject Hydra multirun / sweep syntax.
    """
    for tok in overrides:
        if tok in {"-m", "--multirun"} or "," in tok:
            raise click.UsageError(
                "Parameter sweeps are not supported in `owls run`.\n\n"
                "Use single-value overrides only, e.g.:\n"
                "  owls run Case.toml -- longevity.jack=85 longevity.jill=90\n\n"
                "For sweeps, use the Hydra runner."
            )


def insert_text_as_first_sheet(
    xlsx_path: Path,
    text: str,
    sheet_name: str = "Config (resolved TOML)",
):
    """
    Insert provided text (resolved TOML) as the first worksheet.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.create_sheet(title=sheet_name, index=0)

    for row_idx, line in enumerate(text.splitlines(), start=1):
        ws.cell(row=row_idx, column=1, value=line)

    wb.save(xlsx_path)


def validate_toml(ctx, param, value: Path):
    if value is None:
        return None

    if value.suffix == "":
        value = value.with_suffix(".toml")

    if value.suffix.lower() != ".toml":
        raise click.BadParameter("File must have a .toml extension")

    if not value.exists():
        raise click.BadParameter(f"File '{value}' does not exist")

    if not value.is_file():
        raise click.BadParameter(f"'{value}' is not a file")

    return value


# ---------------------------------------------------------------------
# CLI command
# ---------------------------------------------------------------------


@click.command(
    name="run_hydra",
    context_settings=dict(
        ignore_unknown_options=True,
        allow_extra_args=True,
    ),
)
@click.argument(
    "filename",
    type=click.Path(exists=False, dir_okay=False, path_type=Path),
    callback=validate_toml,
)
@click.pass_context
def cmd_run(ctx, filename: Path):
    """
    Run the solver for an input OWL plan file.

    - The TOML file defines the base case
    - Hydra overrides are applied to TOML BEFORE execution
    - The resolved TOML is embedded in the output workbook
    - Parameter sweeps are NOT supported here
    """

    logger.debug("Executing run command with file: {}", filename)
    logger.debug("Hydra overrides received: {}", ctx.args)

    # -----------------------------
    # Reject sweep syntax
    # -----------------------------
    reject_sweeps(ctx.args)
    # -----------------------------
    # Load Hydra configuration
    # -----------------------------
    load_hydra_cfg(overrides=ctx.args)

    # -----------------------------
    # Extract EXACT Hydra overrides
    # -----------------------------
    logger.info("overrides: {}", ctx.args)
    overrides = hydra_overrides_to_dict(ctx.args)

    logger.info("Scenario overrides: {}", overrides)

    # -----------------------------
    # Output filename
    # -----------------------------
    # Keep filename stable; scenario details live in:
    #   - embedded resolved TOML
    #   - Hydra metadata
    output_filename = filename.with_suffix(".xlsx")

    # -----------------------------
    # Run OWL via core runner
    # -----------------------------
    result = run_single_case(
        case_file=str(filename),
        overrides=overrides,
        output_file=str(output_filename),
    )

    click.echo(f"Case status: {result.status}")

    if result.status != "solved":
        return

    # -----------------------------
    # Insert resolved TOML
    # -----------------------------
    if result.adjusted_toml:
        insert_text_as_first_sheet(
            xlsx_path=output_filename,
            text=result.adjusted_toml,
        )

    click.echo(f"Results saved to: {output_filename}")
